# -*- coding:utf8 -*-

import os, time, json,re, io,sys,codecs,chardet
#reload(sys)
#sys.setdefaultencoding('utf8')
import openpyxl as xls
import openpyxl.workbook
import openpyxl.worksheet
import openpyxl.cell.cell
import openpyxl.styles as xlStyle
from openpyxl.styles import Font, colors, Alignment, PatternFill
import configparser

# directory paths
gram_path = os.getcwd()
root_path = os.path.abspath(os.path.join(gram_path, '../'))
out_path = root_path + '/out'

# parser config.ini
config = configparser.ConfigParser()
config_path_name = root_path + '/config.ini'
config.read(config_path_name, encoding='utf-8')

# Number of single extended statements
n = int(eval(config.get('Page_Selected_Info', 'page3_selected'))[0])


def root_type_list():
    """
    get all the info ralation into all types
    :return:
    """
    all_file_types = []
#    files = os.listdir(out_path)
    files = eval(config.get('Page_Selected_Info','page1_selected'))
    for item in files:
        e_01 = item
        json_file_name = 'variation_'+item+'.json' 
        xls_file_name = 'variation_' + e_01 + '.xlsx'
        all_file_types.append([e_01, json_file_name, xls_file_name])
    return all_file_types


def __get_json(json_file_name):
    """
    get the content from each one json file
    :return:
    """
    json_path_name = out_path + '/' + json_file_name
    json_data = open(json_path_name,'r',encoding='utf8')
    result = json.load(json_data)
    json_data.close()
    return result


def merger_index(index_list):
    """
    :return:index range of the merged columns
    """
    merger_index = []
    root = 2
    d_merger_root = 'A' + str(root)
    merger_start = d_merger_root
    for item in index_list:
        __end_numb = root + item - 1
        merger_end = 'A' + str(__end_numb)
        merger_index.append([merger_start, merger_end])
        root = __end_numb + 1
        merger_start = 'A' + str(root)
    return merger_index


def _judge_file_exist():
    result_file = config.get('zch_test', 'result_file_name')
    file_path_name = out_path + '/' + result_file
    files = os.listdir(out_path)
    if result_file in files:
        os.remove(file_path_name)


def write_xls(file_info):
    # create the workBook
    wb = xls.Workbook()

    # select one worksheet
    ws1 = wb.active

    # update the name of sheet
    ws1.title = file_info[0]

    # prepare the content what you want
    
    middle_file_name = file_info[1]
    root_data = __get_json(middle_file_name)

    # column name writed 
    Titles = eval(config.get('xlsx_info', 'Titles'))
    ws1['B1'] = Titles.get('type')
    ws1['C1'] = Titles.get('number')
    ws1['D1'] = Titles.get('represent')
    ws1['A1'] = Titles.get('goal')

    # Dead work(start line number of content)
    _index_domain = _index_type = _index_No = _index_utt = 2
    # number of rows occupied by each goal
    d_length = 0
    d_length_list = []

    d_name_list = []
    for domain in root_data:

       # _one = list(domain.keys())
        _d_level = list(domain.values())
        domain_keys = list(domain.keys())
        if 'mianCase' in domain_keys:
            domain_keys.remove('mainCase')
        d_name = domain.get(domain_keys[0])
        types = domain.get('mainCase')  
        d_name_list.append(d_name)

        utt_num = 0
        t_name_list = []
        for type_each in types:
            #_t_level = list(type_each.values())
            t_name = type_each.get('name')
            utts =type_each.get('variation')
            t_name_list.append(t_name)

            utt_num = len(utts)
            for number in range(utt_num):
                # No. data writting table
                _class_No = 'C' + str(_index_No)
                ws1[_class_No] = number + 1
                _index_No += 1

                _class_utt = 'D' + str(_index_utt)
                # utt data writting table
                tmp_value_01 = next(utts.__iter__())
                ws1[_class_utt] = tmp_value_01
                _index_utt += 1
                utts.pop(0)

        type_num = len(t_name_list)
        for number in range(type_num):
            _class_type = 'B' + str(_index_type)
            # utt data writting table
            tmp_value_01 = next(t_name_list.__iter__())
            ws1[_class_type] = tmp_value_01
            _index_type += utt_num
            t_name_list.pop(0)

        _class_domain = 'A' + str(_index_domain)
        ws1[_class_domain] = d_name
        d_length = type_num * n
        d_length_list.append(d_length)
        _index_domain += (n * type_num)

    # update the style
    # 0) base patternfill
    backColor_a1_a2 = PatternFill(fill_type='solid', start_color='a6a6a6', end_color='a6a6a6')
    backColor_a3_a4 = PatternFill(fill_type='solid', start_color='c0504d', end_color='c0504d')

    # 1) Set the hight of first row
    row1 = ws1.row_dimensions[1]
    row1.ht = 15

    # 2) Set the width of columns involved
    col01 = ws1.column_dimensions['A']
    col02 = ws1.column_dimensions['B']
    col03 = ws1.column_dimensions['C']
    col04 = ws1.column_dimensions['D']
    col01.width = 28
    col02.width = 35
    col03.width = 10
    col04.width = 60

    # 3) Merger goal column
    merger_s = merger_index(d_length_list)
    for item in merger_s:
        cell = ws1[item[0]]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        item = ':'.join(item)
        ws1.merge_cells(item)

    # 4) title background
    cell_a1 = ws1['A1']
    cell_b1 = ws1['B1']
    cell_c1 = ws1['C1']
    cell_d1 = ws1['D1']
    cell_a1.fill = cell_b1.fill = backColor_a1_a2
    cell_c1.fill = cell_d1.fill = backColor_a3_a4

    # 
    _judge_file_exist()

    # save the xlsx file  
    __result_file = out_path + '/' + file_info[2]
    wb.save(filename=__result_file)


if __name__ == '__main__':
    files_types = root_type_list()
    for item in files_types:
        write_xls(item)
    print('json_to_xlsx.py excution is completed!')
    #tmp_item = ['music','variation_music.json','variation_music_tmp_zch.xlsx']
    #write_xls(tmp_item)

